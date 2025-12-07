import time
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from datetime import datetime
from openpyxl import load_workbook
import unicodedata
from django.db.models import Q
from django.core.validators import validate_email
from django.core.exceptions import ValidationError

from .serializers import *
from .models import *
from apps.audit_logs.models import AuditLog
from common.response_handler import APIResponse
from apps.audit_logs.utils.audit import create_audit_log
from apps.notifications.services.email_service import (
    send_email_updated_notification,
    send_password_changed_notification
)

REQUIRED_COLUMNS = {
    'dni': ['dni'],
    'last_name': ['apellidos', 'apellido', 'last name'],
    'first_name': ['nombres', 'nombre', 'first name'],
    'start_date': ['fecha inicio', 'fechainicio'],
    'position': ['nombre de cargo', 'cargo', 'position', 'NombreCargo'],
    'description': ['descripcion', 'description'],
    'condition': ['condicion', 'condition'],
    'category': ['categoria', 'category'],
    'regimen': ['regimen', 'regime'],
    'identification_code': ['codigo de identificacion', 'codigo', 'identification code', 'CodigoIdentificacion'],
    'role': ['tipo', 'rol', 'role'],
    'descriptionSP': ['descripcionSP', 'descripcion sp', 'descripcionsistema'],
    'end_date': ['fecha fin', 'end date'],
    'resigned_date': ['fecha renuncia', 'fecha de renuncia', 'resigned date'],
    'resigned': ['con renuncia', 'resigned'],
    'establishment': ['establecimiento', 'establishment']
}

OPTIONAL_COLUMNS = {
    'email': ['email', 'correo', 'correo electronico']
}

WORK_DETAILS_COLUMNS = {
    "dni": ["DNI"],
    "worked_days": ["DiasTrabajados"],
    "non_worked_days": ["DiasNoTrabajados"],
    "worked_hours": ["HorasTrabajados"],
    "discount_academic_hours": ["DescuentoHorasAcademicas"],
    "discount_lateness": ["DescuentoTardanzas"],
    "personal_leave_hours": ["PermisoParticular"],
    "sunday_discount": ["DescuentoDominical"],
    "vacation_days": ["DiasVacaciones"],
    "vacation_hours": ["HorasVacaciones"]
}

def normalize(s):
    """Quita acentos, espacios y pasa a minúsculas"""
    if not s:
        return ''
    s = ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
    return s.replace(' ', '').lower()

def to_upper(val):
    return str(val).upper() if val else None

def parse_date(val):
    if not val:
        return None
    if isinstance(val, datetime):
        return val.date() 
    try:
        return datetime.strptime(val, "%d/%m/%Y").date()
    except Exception:
        return None 
    
class ProfileViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated]

    @action(detail=False, methods=['post'], url_path='upload-users')
    def upload_users(self, request):
        start_time = time.time() 
        if not request.user.profile.role == 'admin':
            return Response(
                APIResponse.error(
                    message="No tiene permisos para realizar esta acción.",
                    code=status.HTTP_403_FORBIDDEN
                ),
                status=status.HTTP_403_FORBIDDEN
            )

        file = request.FILES.get('file')
        if not file:
            return Response(
                APIResponse.error(
                    message="No se ha enviado ningún archivo.",
                    code=status.HTTP_400_BAD_REQUEST
                ),
                status=status.HTTP_400_BAD_REQUEST
            )

        if not file.name.endswith('.xlsx'):
            return Response(
                APIResponse.error(
                    message="El archivo debe ser un Excel (.xlsx).",
                    code=status.HTTP_400_BAD_REQUEST
                ),
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            workbook = load_workbook(filename=file)
            ws = workbook.active
        except Exception as e:
            return Response(
                APIResponse.error(
                    message=f"Error al abrir el archivo: {str(e)}",
                    code=status.HTTP_400_BAD_REQUEST
                ),
                status=status.HTTP_400_BAD_REQUEST
            )

        headers = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]
        normalized_headers = [normalize(h) for h in headers]
        column_map = {}

        for idx, nh in enumerate(normalized_headers):
            for field, variants in REQUIRED_COLUMNS.items():
                if nh in [normalize(v) for v in variants]:
                    column_map[idx] = field
        
        for idx, nh in enumerate(normalized_headers):
            for field, variants in OPTIONAL_COLUMNS.items():
                if nh in [normalize(v) for v in variants]:
                    column_map[idx] = field

        missing = [field for field in REQUIRED_COLUMNS.keys() if field not in column_map.values()]
        if missing:
            return Response(
                APIResponse.error(
                    message=f"Faltan columnas obligatorias en el Excel: {', '.join(missing)}",
                    code=status.HTTP_400_BAD_REQUEST
                ),
                status=status.HTTP_400_BAD_REQUEST
            )
        
        results = {
            'created_count': 0,
            'updated_count': 0,
            'skipped_rows': 0
        }
        error_messages = [] 

        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            row_data = {column_map[idx]: cell.value for idx, cell in enumerate(row) if idx in column_map}

            dni = to_upper(row_data.get('dni'))
            last_name = to_upper(row_data.get('last_name'))
            first_name = to_upper(row_data.get('first_name'))

            if not dni or not last_name or not first_name:
                results['skipped_rows'] += 1
                error_messages.append(f"Fila {row_idx}: DNI, Apellido o Nombre faltantes. (DNI: {dni or 'N/A'})")
                continue

            email = row_data.get('email')
            user_data = {'first_name': first_name, 'last_name': last_name}
            if email:
                user_data['email'] = email

            try:
                user, created = User.objects.update_or_create(
                    username=dni,
                    defaults=user_data
                )
                if created:
                    user.set_password(dni)
                    user.save()
                    results['created_count'] += 1
                else:
                    results['updated_count'] += 1
            except Exception as e:
                results['skipped_rows'] += 1
                error_messages.append(f"Fila {row_idx}: Error al crear/actualizar usuario con DNI {dni}: {str(e)}")
                continue

            resigned_value = row_data.get('resigned')
            resigned_bool = resigned_value in [1, "1", True, "TRUE", "True"]
            profile_data = {
                'user': user, 
                'dni': dni, 
                'role': 'user',
                'position': to_upper(row_data.get('position')),
                'description': to_upper(row_data.get('description')), 
                'descriptionSP': to_upper(row_data.get('descriptionSP')),
                'start_date': parse_date(row_data.get('start_date')),
                'end_date': parse_date(row_data.get('end_date')), 
                'resigned_date': parse_date(row_data.get('resigned_date')),
                'resigned': resigned_bool, 
                'regimen': to_upper(row_data.get('regimen')), 
                'category': to_upper(row_data.get('category')),
                'condition': to_upper(row_data.get('condition')),
                'identification_code': to_upper(row_data.get('identification_code')),
                'establishment': to_upper(row_data.get('establishment')),
            }
            try:
                Profile.objects.update_or_create(user=user, defaults=profile_data)
            except Exception as e:
                error_messages.append(f"Fila {row_idx}: Error al actualizar perfil con DNI {dni}: {str(e)}")


        final_messages = []
        
        if results['created_count'] > 0:
            final_messages.append(f"{results['created_count']} usuarios nuevos creados.")
        
        if results['updated_count'] > 0:
            final_messages.append(f"{results['updated_count']} usuarios actualizados.")

        if results['skipped_rows'] > 0:
            final_messages.append(f"{results['skipped_rows']} filas fueron saltadas o con errores. ({len(error_messages)} errores detallados).")

        final_messages.extend(error_messages)

        if not final_messages:
            final_messages.append("Procesamiento finalizado sin cambios visibles o errores.")
        
        main_message = "Procesamiento de carga de usuarios finalizado."

        description_text = "\n".join(final_messages)
        AuditLog.objects.create(
            profile=request.user.profile,
            action="CARGA DE USUARIOS",
            description=description_text
        )

        return Response(
            APIResponse.success(
                message=main_message, 
                data={
                    'messages': final_messages, 
                    'created_count': results['created_count'],
                    'updated_count': results['updated_count'],
                    'skipped_rows': results['skipped_rows']
                },
                meta={  
                    "durationMs": int((time.time() - start_time) * 1000)
                }
            ),
            status=status.HTTP_200_OK
        )
    
    @action(detail=False, methods=['post'], url_path='upload-work-details')
    def upload_work_details(self, request):
        start_time = time.time() 

        if not request.user.profile.role == 'admin':
            return Response(
                APIResponse.error(
                    message="No tiene permisos para realizar esta acción.",
                    code=status.HTTP_403_FORBIDDEN
                ),
                status=status.HTTP_403_FORBIDDEN
            )

        file = request.FILES.get('file')
        if not file:
            return Response(
                APIResponse.error(
                    message="No se ha enviado ningún archivo.",
                    code=status.HTTP_400_BAD_REQUEST
                ),
                status=status.HTTP_400_BAD_REQUEST
            )

        if not file.name.endswith('.xlsx'):
            return Response(
                APIResponse.error(
                    message="El archivo debe ser un Excel (.xlsx).",
                    code=status.HTTP_400_BAD_REQUEST
                ),
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            workbook = load_workbook(filename=file)
            ws = workbook.active
        except Exception as e:
            return Response(
                APIResponse.error(
                    message=f"Error al abrir el archivo: {str(e)}",
                    code=status.HTTP_400_BAD_REQUEST
                ),
                status=status.HTTP_400_BAD_REQUEST
            )

        headers = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]
        normalized_headers = [normalize(h) for h in headers]

        column_map = {}
        for idx, nh in enumerate(normalized_headers):
            for field, variants in WORK_DETAILS_COLUMNS.items():
                if nh in [normalize(v) for v in variants]:
                    column_map[idx] = field

        missing = [field for field in WORK_DETAILS_COLUMNS.keys() if field not in column_map.values()]
        if missing:
            return Response(
                APIResponse.error(
                    message=f"Faltan columnas obligatorias en el Excel: {', '.join(missing)}",
                    code=status.HTTP_400_BAD_REQUEST
                ),
                status=status.HTTP_400_BAD_REQUEST
            )

        created_count = 0
        updated_count = 0
        skipped_count = 0
        detailed_messages = [] 
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            row_data = {column_map[idx]: cell.value for idx, cell in enumerate(row) if idx in column_map}

            dni = to_upper(row_data.get('dni'))

            if not dni:
                skipped_count += 1
                detailed_messages.append(f"Fila {row_idx}: DNI es obligatorio.")
                continue

            try:
                profile = Profile.objects.get(dni=dni)
            except Profile.DoesNotExist:
                skipped_count += 1
                detailed_messages.append(f"Fila {row_idx}: No existe Profile con DNI {dni}.")
                continue
            except Exception as e:
                skipped_count += 1
                detailed_messages.append(f"Fila {row_idx}: Error al buscar Profile con DNI {dni}: {str(e)}")
                continue
            
            try:
                work_data = {
                    'worked_days': int(row_data.get('worked_days') or 0),
                    'non_worked_days': int(row_data.get('non_worked_days') or 0),
                    'worked_hours': int(row_data.get('worked_hours') or 0),
                    'discount_academic_hours': int(row_data.get('discount_academic_hours') or 0),
                    'discount_lateness': int(row_data.get('discount_lateness') or 0),
                    'personal_leave_hours': int(row_data.get('personal_leave_hours') or 0),
                    'sunday_discount': int(row_data.get('sunday_discount') or 0),
                    'vacation_days': int(row_data.get('vacation_days') or 0),
                    'vacation_hours': int(row_data.get('vacation_hours') or 0),
                }
            except ValueError as e:
                skipped_count += 1
                detailed_messages.append(f"Fila {row_idx}: Error de formato numérico en datos laborales: {str(e)}")
                continue
            except Exception as e:
                skipped_count += 1
                detailed_messages.append(f"Fila {row_idx}: Error al preparar datos laborales: {str(e)}")
                continue

            try:
                work_details, created = ProfileWorkDetails.objects.update_or_create(
                    profile=profile,
                    defaults=work_data
                )
            except Exception as e:
                skipped_count += 1
                detailed_messages.append(f"Fila {row_idx}: Error al crear/actualizar WorkDetails para DNI {dni}: {str(e)}")
                continue

            if created:
                created_count += 1
            else:
                updated_count += 1

        final_messages = []
        
        if created_count > 0:
            final_messages.append(f"{created_count} detalles de usuarios fueron creados.")
        
        if updated_count > 0:
            final_messages.append(f"{updated_count} detalles de usuarios fueron actualizados.")

        if skipped_count > 0:
            final_messages.append(f"{skipped_count} filas fueron saltadas o con errores. ({len(detailed_messages)} errores detallados).")

        final_messages.extend(detailed_messages)

        if not final_messages:
            main_message = "Procesamiento de Work Details finalizado sin cambios visibles o errores."
        else:
            main_message = "Procesamiento de Work Details finalizado."

        description_text = "\n".join(final_messages)
        AuditLog.objects.create(
            profile=request.user.profile,
            action="CARGA DE WORK DETAILS",
            description=description_text
        )

        return Response(
            APIResponse.success(
                message=main_message, 
                data={
                    'created_count': created_count,
                    'updated_count': updated_count,
                    'skipped_count': skipped_count,
                    'messages': final_messages, 
                },
                meta={
                    "durationMs": int((time.time() - start_time) * 1000)
                }
            ),
            status=status.HTTP_200_OK
        )
    @action(detail=False, methods=['get'], url_path='list-users')
    def list_users(self, request):
        if not request.user.profile.role == 'admin':
            return Response(
                APIResponse.error(
                    message="No tiene permisos para realizar esta acción.",
                    code=status.HTTP_403_FORBIDDEN
                ),
                status=status.HTTP_403_FORBIDDEN
            )

        try:
            page = int(request.query_params.get('page', 1))
            page_size = int(request.query_params.get('page_size', 20))
            if page < 1:
                page = 1
            if page_size < 1:
                page_size = 20
        except ValueError:
            page = 1
            page_size = 20

        search = request.query_params.get('search', '').strip()

        offset = (page - 1) * page_size
        limit = offset + page_size

        queryset = Profile.objects.select_related('user').filter(role='user').order_by('-created_at')

        if search:
            queryset = queryset.filter(
                Q(user__first_name__icontains=search) |
                Q(user__last_name__icontains=search)
            )

        total = queryset.count()
        profiles = queryset[offset:limit]

        results = []
        for p in profiles:
            results.append({
                "id": str(p.id),
                "dni": p.dni,
                "first_name": p.user.first_name if p.user else None,
                "last_name": p.user.last_name if p.user else None,
                "email": p.user.email if p.user else None,
                "role": p.role,
                "condition": p.condition,
                "position": p.position,
                "created_at": p.created_at.isoformat() if p.created_at else None,
            })

        pagination = {
            "current_page": page,
            "page_size": page_size,
            "total_items": total,
            "total_pages": (total + page_size - 1) // page_size,
            "has_next": limit < total,
            "has_previous": page > 1
        }

        return Response(
            APIResponse.success(
                data=results,
                message=f"{len(results)} usuarios obtenidos.",
                meta={"pagination": pagination}
            ),
            status=status.HTTP_200_OK
        )


    @action(detail=False, methods=['get'], url_path='me')
    def me(self, request):
        profile = getattr(request.user, 'profile', None)
        if not profile:
            return Response(
                APIResponse.error(
                    message="No se encontró el perfil del usuario.",
                    code=status.HTTP_404_NOT_FOUND
                ),
                status=status.HTTP_404_NOT_FOUND
            )

        work_details = getattr(profile, 'work_details', None)

        data = {
            "id": str(profile.id),
            "dni": profile.dni,
            "role": profile.role,
            "first_name": request.user.first_name,
            "last_name": request.user.last_name,
            "email": request.user.email,
            "username": request.user.username,
            "position": profile.position,
            "description": profile.description,
            "descriptionSP": profile.descriptionSP,
            "start_date": profile.start_date.isoformat() if profile.start_date else None,
            "end_date": profile.end_date.isoformat() if profile.end_date else None,
            "resigned_date": profile.resigned_date.isoformat() if profile.resigned_date else None,
            "resigned": profile.resigned,
            "regimen": profile.regimen,
            "category": profile.category,
            "condition": profile.condition,
            "identification_code": profile.identification_code,
            "establishment": profile.establishment,
            "is_active": profile.is_active,
            "work_details": {
                "worked_days": work_details.worked_days if work_details else 0,
                "non_worked_days": work_details.non_worked_days if work_details else 0,
                "worked_hours": work_details.worked_hours if work_details else 0,
                "discount_academic_hours": work_details.discount_academic_hours if work_details else 0,
                "discount_lateness": work_details.discount_lateness if work_details else 0,
                "personal_leave_hours": work_details.personal_leave_hours if work_details else 0,
                "sunday_discount": work_details.sunday_discount if work_details else 0,
                "vacation_days": work_details.vacation_days if work_details else 0,
                "vacation_hours": work_details.vacation_hours if work_details else 0,
            } if work_details else None
        }

        return Response(
            APIResponse.success(
                data=data,
                message="Perfil obtenido correctamente."
            ),
            status=status.HTTP_200_OK
        )
    
    @action(detail=False, methods=['patch'], url_path='update-email')
    def update_email(self, request):
        user = request.user
        new_email = request.data.get("email")

        if not new_email:
            return Response(
                APIResponse.error(
                    message="Debe proporcionar un correo electrónico.",
                    code=status.HTTP_400_BAD_REQUEST
                ),
                status=status.HTTP_400_BAD_REQUEST
            )
            
        new_email = new_email.strip().lower()

        try:
            validate_email(new_email)
        except ValidationError:
            return Response(
                APIResponse.error(
                    message="El formato del correo electrónico es inválido.",
                    code=status.HTTP_400_BAD_REQUEST
                ),
                status=status.HTTP_400_BAD_REQUEST
            )

        if user.email == new_email:
            return Response(
                APIResponse.error(
                    message="El correo electrónico ingresado es igual al actual.",
                    code=status.HTTP_400_BAD_REQUEST
                ),
                status=status.HTTP_400_BAD_REQUEST
            )

        if User.objects.exclude(id=user.id).filter(email=new_email).exists():
            return Response(
                APIResponse.error(
                    message="El correo electrónico ya está siendo usado por otro usuario.",
                    code=status.HTTP_409_CONFLICT
                ),
                status=status.HTTP_409_CONFLICT
            )

        old_email = user.email 

        user.email = new_email
        user.save(update_fields=["email"])

        create_audit_log(
            user=user,
            action="UPDATE_EMAIL",
            description="El usuario actualizó su dirección de correo.",
            extra_data={
                "old_email": old_email,
                "new_email": new_email
            }
        )

        send_email_updated_notification(user, new_email)

        return Response(
            APIResponse.success(
                data={"email": new_email},
                message="Correo electrónico actualizado correctamente."
            ),
            status=status.HTTP_200_OK
        )

    
    @action(detail=False, methods=['post'], url_path='change-password')
    def change_password(self, request):
        serializer = ChangePasswordSerializer(
            data=request.data,
            context={"request": request}
        )

        if not serializer.is_valid():
            return Response(
                APIResponse.error(
                    message="Error al cambiar la contraseña.",
                    code=status.HTTP_400_BAD_REQUEST,
                    errors=serializer.errors
                ),
                status=status.HTTP_400_BAD_REQUEST
            )

        user = request.user
        new_password = serializer.validated_data["new_password"]

        audit_extra = {
            "user_id": user.id,
            "username": user.username,
            "email": user.email,
            "ip": request.META.get("REMOTE_ADDR"),
            "agent": request.headers.get("User-Agent")
        }

        user.set_password(new_password)
        user.save()

        create_audit_log(
            user=user,
            action="CAMBIO DE CONTRASEÑA",
            description="El usuario cambió su contraseña.",
            extra_data=audit_extra
        )

        send_password_changed_notification(user)

        return Response(
            APIResponse.success(
                message="La contraseña se actualizó correctamente."
            ),
            status=status.HTTP_200_OK
        )
