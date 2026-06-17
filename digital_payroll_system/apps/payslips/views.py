import time
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from uuid import UUID
from openpyxl import load_workbook
from apps.profiles.models import Profile
from .models import Payslip
import unicodedata
from common.response_handler import APIResponse
from apps.audit_logs.models import AuditLog
from datetime import datetime
from decimal import Decimal
from django.template.loader import render_to_string
from django.core.files.base import ContentFile
from xhtml2pdf import pisa
from io import BytesIO
from django.utils import timezone
from django.shortcuts import get_object_or_404
from apps.notifications.services.email_service import send_payslip_email
from apps.notifications.services.qr_service import generate_qr_code
from django.db.models import Max, Q, Subquery, OuterRef, F, Value, CharField
from django.db.models.functions import Concat, ExtractMonth, ExtractYear
from django.db import transaction

MONTHS_ES = [
    "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO",
    "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"
]

REQUIRED_PAYSLIP_COLUMNS = {
    "dni": ["DNI"],
    "concept": ["Concepto"],
    "amount": ["Monto"],
    "data_source": ["OrigenDato"],
    "payroll_type": ["TipoPlanilla"],
    "data_type": ["TipoDato"],
    "position_order": ["Posicion"],
    "issue_date": ["Periodo"],
}

def normalize(s):
    """Quita acentos, espacios y pasa a minúsculas"""
    if not s:
        return ''
    s = ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
    return s.replace(' ', '').lower()

def to_upper(val):
    return str(val).upper() if val else None

def parse_date(val):
    if not val:
        return None
    if isinstance(val, datetime):
        return val.date() 
    try:
        return datetime.strptime(val, "%d/%m/%Y").date()
    except Exception:
        return None 

def parse_period(period_text):
    try:
        months = {
            "ENERO": 1, "FEBRERO": 2, "MARZO": 3, "ABRIL": 4,
            "MAYO": 5, "JUNIO": 6, "JULIO": 7, "AGOSTO": 8,
            "SEPTIEMBRE": 9, "OCTUBRE": 10, "NOVIEMBRE": 11, "DICIEMBRE": 12
        }
        parts = period_text.upper().split()
        month = months.get(parts[0], 1)
        year = int(parts[1])
        return datetime(year, month, 1).date()
    except Exception:
        return None
     
class PayslipUploadViewSet(viewsets.ViewSet):
    """
    Upload payslips from Excel file.
    """
    permission_classes = []

    @action(detail=False, methods=['post'], url_path='upload-payslips')
    def upload_payslips(self, request):
        start_time = time.time() 

        if not request.user.profile.role == 'admin':
            return Response(
                APIResponse.error(
                    message="No tiene permisos para realizar esta acción.",
                    code=status.HTTP_403_FORBIDDEN
                ),
                status=status.HTTP_403_FORBIDDEN
            )

        file = request.FILES.get('file')
        if not file:
            return Response(
                APIResponse.error(
                    message="No se ha enviado ningún archivo.",
                    code=status.HTTP_400_BAD_REQUEST
                ),
                status=status.HTTP_400_BAD_REQUEST
            )

        if not file.name.endswith('.xlsx'):
            return Response(
                APIResponse.error(
                    message="El archivo debe ser un Excel (.xlsx).",
                    code=status.HTTP_400_BAD_REQUEST
                ),
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            workbook = load_workbook(filename=file)
            ws = workbook.active
        except Exception as e:
            return Response(
                APIResponse.error(
                    message=f"Error al abrir el archivo: {str(e)}",
                    code=status.HTTP_400_BAD_REQUEST
                ),
                status=status.HTTP_400_BAD_REQUEST
            )
            
        headers = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]
        normalized_headers = [normalize(h) for h in headers]

        column_map = {}
        for idx, nh in enumerate(normalized_headers):
            for field, variants in REQUIRED_PAYSLIP_COLUMNS.items():
                if nh in [normalize(v) for v in variants]:
                    column_map[idx] = field

        missing = [field for field in REQUIRED_PAYSLIP_COLUMNS.keys() if field not in column_map.values()]
        if missing:
            return Response(
                APIResponse.error(
                    message=f"Faltan columnas obligatorias en el Excel: {', '.join(missing)}",
                    code=status.HTTP_400_BAD_REQUEST
                ),
                status=status.HTTP_400_BAD_REQUEST
            )

        created_count = 0
        skipped_count = 0
        error_messages = [] 

        profiles_map = {p.dni: p for p in Profile.objects.all()}

        try:
            with transaction.atomic():
                for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                    row_data = {column_map[idx]: cell.value for idx, cell in enumerate(row) if idx in column_map}

                    dni = str(row_data.get('dni')).strip() if row_data.get('dni') else None
                
                    if not dni:
                        skipped_count += 1
                        error_messages.append(f"Fila {row_idx}: DNI no encontrado. Se saltó la fila.")
                        continue

                    profile = profiles_map.get(dni)
                    if not profile:
                        skipped_count += 1
                        error_messages.append(f"Fila {row_idx}: Usuario con DNI {dni} no existe. Se saltó la fila.")
                        continue

                    period_text = str(row_data.get('issue_date'))
                    issue_date = parse_period(period_text)
                    if not issue_date:
                        skipped_count += 1
                        error_messages.append(f"Fila {row_idx}: Periodo '{period_text}' inválido. Se saltó la fila.")
                        continue
                    
                    try:
                        amount = Decimal(str(row_data.get('amount')))
                    except Exception:
                        skipped_count += 1
                        error_messages.append(f"Fila {row_idx}: Monto inválido '{row_data.get('amount')}'. Se saltó la fila.")
                        continue
                    
                    concept = str(row_data.get('concept')).upper()

                    if Payslip.objects.filter(
                        profile=profile,
                        concept=concept,
                        issue_date__year=issue_date.year,
                        issue_date__month=issue_date.month
                    ).exists():
                        skipped_count += 1
                        error_messages.append(
                            f"Fila {row_idx}: Ya existe una boleta con el concepto '{concept}' "
                            f"para el periodo {issue_date.strftime('%Y-%m')}. Se saltó la fila."
                        )
                        continue
                    
                    try:
                        Payslip.objects.create(
                            profile=profile,
                            concept=concept,
                            amount=amount,
                            data_source=str(row_data.get('data_source')).upper(),
                            payroll_type=str(row_data.get('payroll_type')).upper(),
                            data_type=str(row_data.get('data_type')).upper(),
                            position_order=int(row_data.get('position_order')),
                            issue_date=issue_date,
                            pdf_file='',
                            view_status='unseen'
                        )
                        created_count += 1
                    except Exception as e:
                        skipped_count += 1
                        error_messages.append(f"Fila {row_idx}: Error al crear la boleta para DNI {dni}: {str(e)}. Se saltó la fila.")

        except Exception as e:
            return Response(
                APIResponse.error(
                    message=f"Error crítico en la transacción de la base de datos: {str(e)}",
                    code=status.HTTP_500_INTERNAL_SERVER_ERROR
                ),
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        final_messages = []
        if created_count > 0:
            final_messages.append(f"{created_count} boletas creadas exitosamente.")
        
        if skipped_count > 0:
            final_messages.append(f"{skipped_count} filas fueron saltadas o con errores. ({len(error_messages)} errores detallados).")

        final_messages.extend(error_messages)

        if not final_messages:
            main_message = "Procesamiento de Boletas finalizado sin cambios visibles o errores."
        else:
            main_message = "Procesamiento de Boletas finalizado."

        description_text = "\n".join(final_messages)
        AuditLog.objects.create(
            profile=request.user.profile,
            action="CARGA DE BOLETAS",
            description=description_text
        )

        return Response(
            APIResponse.success(
                message=main_message, 
                data={
                    'messages': final_messages, 
                    'created_count': created_count,
                    'skipped_count': skipped_count
                },
                meta={
                    "durationMs": int((time.time() - start_time) * 1000)
                }
            ),
            status=status.HTTP_200_OK
        )
    
    @action(detail=False, methods=['delete'], url_path='clear-payslips')
    def clear_payslips(self, request):
        if not request.user.profile.role == 'admin':
            return Response(
                APIResponse.error(
                    message="No tiene permisos para realizar esta acción.",
                    code=status.HTTP_403_FORBIDDEN
                ),
                status=status.HTTP_403_FORBIDDEN
            )

        payslips = Payslip.objects.all()
        total_deleted = payslips.count()

        for ps in payslips:
            if ps.pdf_file:
                ps.pdf_file.delete(save=False)

        payslips.delete()

        AuditLog.objects.create(
            profile=request.user.profile,
            action="ELIMINAR BOLETAS",
            description=f"Se eliminaron {total_deleted} boletas de todos los usuarios."
        )

        return Response(
            APIResponse.success(
                message=f"Se eliminaron {total_deleted} boletas correctamente."
            ),
            status=status.HTTP_200_OK
        )
    
    @action(detail=False, methods=['delete'], url_path='delete-payslip')
    def delete_payslip(self, request):
        payslip_id = request.data.get('id')

        if not payslip_id:
            return Response(
                APIResponse.error(
                    message="Se requiere el ID de la boleta para eliminar.",
                    code=status.HTTP_400_BAD_REQUEST
                ),
                status=status.HTTP_400_BAD_REQUEST
            )

        if not request.user.profile.role == 'admin':
            return Response(
                APIResponse.error(
                    message="No tiene permisos para realizar esta acción.",
                    code=status.HTTP_403_FORBIDDEN
                ),
                status=status.HTTP_403_FORBIDDEN
            )

        try:
            uuid_obj = UUID(payslip_id, version=4)
        except (ValueError, TypeError):
            return Response(
                APIResponse.error(
                    message=f"El ID '{payslip_id}' no es un UUID válido.",
                    code=status.HTTP_400_BAD_REQUEST
                ),
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            payslip = Payslip.objects.get(id=payslip_id)
        except Payslip.DoesNotExist:
            return Response(
                APIResponse.error(
                    message=f"No se encontró ninguna boleta con ID {payslip_id}.",
                    code=status.HTTP_404_NOT_FOUND
                ),
                status=status.HTTP_404_NOT_FOUND
            )

        payslip.delete()

        AuditLog.objects.create(
            profile=request.user.profile,
            action="ELIMINAR BOLETA",
            description=f"Se eliminó la boleta del usuario {payslip.profile.dni}."
        )

        return Response(
            APIResponse.success(
                message=f"Boleta eliminada correctamente."
            ),
            status=status.HTTP_200_OK
        )
    
    @action(detail=False, methods=['get'], url_path='list-payslips')
    def list_payslips(self, request):
        """
        Vista para el Administrador: Lista una única fila por usuario y periodo (mes/año),
        mostrando los montos totales agregados de Ingresos, Descuentos y el Neto Líquido.
        """
        if not request.user.profile.role == 'admin':
            return Response(
                APIResponse.error(message="No tiene permisos para realizar esta acción.", code=status.HTTP_403_FORBIDDEN),
                status=status.HTTP_403_FORBIDDEN
            )

        try:
            page = int(request.query_params.get('page', 1))
            page_size = int(request.query_params.get('page_size', 20))
        except ValueError:
            page = 1
            page_size = 20
        
        page = max(page, 1)
        page_size = max(page_size, 1)

        dni = request.query_params.get('dni')
        name = request.query_params.get('name')
        status_view = request.query_params.get('status')
        month = request.query_params.get('month')
        year = request.query_params.get('year')
        base_queryset = Payslip.objects.select_related('profile', 'profile__user')

        if dni:
            base_queryset = base_queryset.filter(profile__dni__icontains=dni)

        base_queryset = base_queryset.annotate(
            full_name_concat=Concat(
                F('profile__user__first_name'),
                Value(' '),
                F('profile__user__last_name'),
                output_field=CharField()
            )
        )

        if name:
            tokens = [t.strip() for t in name.split() if t.strip()]
            for token in tokens: 
                base_queryset = base_queryset.filter(full_name_concat__icontains=token)

        if status_view:
            base_queryset = base_queryset.filter(view_status=status_view)

        if month:
            try: base_queryset = base_queryset.filter(issue_date__month=int(month))
            except: pass

        if year:
            try: base_queryset = base_queryset.filter(issue_date__year=int(year))
            except: pass

        grouped_qs = base_queryset.values('profile', 'issue_date').annotate(
            total_ingresos=Max('amount', filter=Q(data_source='TOTALINGRESOS')),
            total_descuentos=Max('amount', filter=Q(data_source='TOTALDSCTO')),
            liquido_pagar=Max('amount', filter=Q(data_source='LIQUIDOPAGAR')),
            reference_id=Max('id'),
            has_pdf=Max('pdf_file', filter=Q(pdf_file__isnull=False, pdf_file__ne=''))
        ).order_by('-issue_date')

        total = grouped_qs.count()
        offset = (page - 1) * page_size
        paginated_groups = grouped_qs[offset: offset + page_size]

        results = []
        for g in paginated_groups:
            ref_payslip = Payslip.objects.select_related('profile', 'profile__user').filter(
                profile_id=g['profile'], 
                issue_date=g['issue_date']
            ).first()
            
            if not ref_payslip:
                continue

            user = ref_payslip.profile.user
            full_name = f"{user.first_name} {user.last_name}".strip() if user else None
            
            try:
                pdf_url = request.build_absolute_uri(ref_payslip.pdf_file.url) if ref_payslip.pdf_file else None
            except Exception:
                pdf_url = None

            month_idx = g['issue_date'].month
            month_name = MONTHS_ES[month_idx - 1]

            results.append({
                "id": str(g['reference_id']), 
                "profile_id": str(ref_payslip.profile.id),
                "profile_dni": ref_payslip.profile.dni,
                "full_name": full_name,
                "issue_date": g['issue_date'].isoformat(),
                "period_es": f"{month_name} {g['issue_date'].year}",
                "view_status": 'generated' if g['has_pdf'] else 'unseen',
                "concept": "BOLETA RESUMEN MENSUAL",
                "total_ingresos": float(g['total_ingresos'] or 0.00),
                "total_descuentos": float(g['total_descuentos'] or 0.00),
                "amount": float(g['liquido_pagar'] or 0.00),
                "pdf_url": pdf_url,
            })

        pagination = {
            "current_page": page,
            "page_size": page_size,
            "total_items": total,
            "total_pages": (total + page_size - 1) // page_size,
            "has_next": offset + page_size < total,
            "has_previous": page > 1
        }

        return Response(
            APIResponse.success(data=results, message=f"{len(results)} periodos de boletas obtenidos.", meta={"pagination": pagination}),
            status=status.HTTP_200_OK
        )

    @action(detail=False, methods=['get'], url_path='my-payslips')
    def my_payslips(self, request):
        """
        Vista para el Usuario Común: Lista un único registro resumen por cada mes,
        evitando que visualice filas repetidas de sus propios conceptos.
        """
        user = request.user
        if user.profile.role != 'user':
            return Response(
                APIResponse.error(message="No tiene permisos para acceder a sus boletas.", code=status.HTTP_403_FORBIDDEN),
                status=status.HTTP_403_FORBIDDEN
            )

        profile = user.profile
        month = request.query_params.get('month') 
        year = request.query_params.get('year') 
        payslips_qs = Payslip.objects.filter(profile=profile)

        if year:
            try: payslips_qs = payslips_qs.filter(issue_date__year=int(year))
            except ValueError: return Response(APIResponse.error(message="Año inválido"), status=400)

        if month:
            try:
                m_int = int(month)
                if not 1 <= m_int <= 12: raise ValueError
                payslips_qs = payslips_qs.filter(issue_date__month=m_int)
            except ValueError: return Response(APIResponse.error(message="Mes inválido"), status=400)

        grouped_qs = payslips_qs.values('issue_date').annotate(
            total_ingresos=Max('amount', filter=Q(data_source='TOTALINGRESOS')),
            total_descuentos=Max('amount', filter=Q(data_source='TOTALDSCTO')),
            liquido_pagar=Max('amount', filter=Q(data_source='LIQUIDOPAGAR')),
            reference_id=Max('id'),
            has_pdf=Max('pdf_file', filter=Q(pdf_file__isnull=False, pdf_file__ne=''))
        ).order_by('-issue_date')

        page = int(request.query_params.get('page', 1))
        page_size = 20
        offset = (page - 1) * page_size
        limit = offset + page_size

        total = grouped_qs.count()
        paginated_groups = grouped_qs[offset:limit]

        results = []
        for g in paginated_groups:
            ref_payslip = Payslip.objects.filter(id=g['reference_id']).first()
            if not ref_payslip:
                continue

            try:
                pdf_url = request.build_absolute_uri(ref_payslip.pdf_file.url) if ref_payslip.pdf_file else None
            except Exception:
                pdf_url = None

            month_idx = g['issue_date'].month
            month_name = MONTHS_ES[month_idx - 1]

            results.append({
                "id": str(g['reference_id']), 
                "profile_id": str(profile.id),
                "profile_dni": profile.dni,
                "issue_date": g['issue_date'].isoformat(),
                "period_es": f"{month_name} {g['issue_date'].year}",
                "view_status": 'generated' if g['has_pdf'] else 'unseen',
                "concept": "BOLETA DE PAGO MENSUAL",
                "total_ingresos": float(g['total_ingresos'] or 0.00),
                "total_descuentos": float(g['total_descuentos'] or 0.00),
                "amount": float(g['liquido_pagar'] or 0.00), 
                "pdf_url": pdf_url,
            })

        pagination = {
            "current_page": page,
            "page_size": page_size,
            "total_items": total,
            "total_pages": (total + page_size - 1) // page_size,
            "has_next": limit < total,
            "has_previous": page > 1
        }

        return Response(
            APIResponse.success(data=results, message=f"{len(results)} boletas mensuales obtenidas.", meta={"pagination": pagination}),
            status=status.HTTP_200_OK
        )
    
    @action(detail=False, methods=['get'], url_path='view-payslip')
    def view_payslip(self, request):
        payslip_id = request.query_params.get('id')
        if not payslip_id:
            return Response(
                APIResponse.error(
                    message="Debe proporcionar el ID de la boleta.",
                    code=status.HTTP_400_BAD_REQUEST
                ),
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            UUID(payslip_id, version=4)
        except ValueError:
            return Response(
                APIResponse.error(
                    message="El ID proporcionado no es un UUID válido.",
                    code=status.HTTP_400_BAD_REQUEST
                ),
                status=status.HTTP_400_BAD_REQUEST
            )

        payslip = get_object_or_404(Payslip, id=payslip_id)

        is_admin = request.user.profile.role == 'admin'
        is_owner = payslip.profile.user == request.user

        if not is_owner and not is_admin:
            return Response(
                APIResponse.error(
                    message="No tiene permiso para acceder a esta boleta.",
                    code=status.HTTP_403_FORBIDDEN
                ),
                status=status.HTTP_403_FORBIDDEN
            )

        if not payslip.pdf_file:
            return Response(
                APIResponse.error(
                    message="La boleta no tiene un archivo PDF asociado.",
                    code=status.HTTP_400_BAD_REQUEST
                ),
                status=status.HTTP_400_BAD_REQUEST
            )
        if payslip.view_status != 'seen':
            payslip.view_status = 'seen'
            payslip.save()

        if is_admin and not is_owner:
            description = (
                f"El administrador {request.user.username} visualizó y marcó como vista "
                f"la boleta {payslip.id} perteneciente al usuario {payslip.profile.user.username} "
                f"del periodo {payslip.issue_date}."
            )
        else:
            description = (
                f"El usuario {request.user.username} visualizó la boleta "
                f"{payslip.id} del periodo {payslip.issue_date}."
            )

        AuditLog.objects.create(
            profile=request.user.profile,
            action="VISUALIZAR BOLETA",
            description=description
        )

        return Response(
            APIResponse.success(
                message="Boleta visualizada correctamente.",
                data={
                    "payslip_id": str(payslip.id),
                    "pdf_url": request.build_absolute_uri(payslip.pdf_file.url),
                    "status": payslip.view_status,
                    "issue_date": payslip.issue_date,
                    "concept": payslip.concept,
                    "amount": payslip.amount
                }
            ),
            status=status.HTTP_200_OK
        )

    @action(detail=False, methods=['get'], url_path='generate-payslip')
    def generate_payslip(self, request):
        user = request.user
        profile = user.profile

        is_admin = profile.role == 'admin'
        is_user = profile.role == 'user'

        if not (is_admin or is_user):
            return Response(
                APIResponse.error(message="No tiene permisos para generar boletas.", code=status.HTTP_403_FORBIDDEN),
                status=status.HTTP_403_FORBIDDEN
            )

        reference_id = request.query_params.get('id')
        if not reference_id:
            return Response(
                APIResponse.error(message="Debe proporcionar el parámetro 'id' de la boleta."),
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            if is_admin:
                reference_payslip = Payslip.objects.get(id=reference_id)
            else:
                reference_payslip = Payslip.objects.get(id=reference_id, profile=profile)
        except Payslip.DoesNotExist:
            return Response(
                APIResponse.error(message="Boleta no encontrada o no tiene permiso."),
                status=status.HTTP_404_NOT_FOUND
            )
            
        payslip_owner_profile = reference_payslip.profile
        payslip_owner_user = payslip_owner_profile.user
        target_date = reference_payslip.issue_date

        all_concepts = Payslip.objects.filter(
            profile=payslip_owner_profile,
            issue_date__year=target_date.year,
            issue_date__month=target_date.month
        ).order_by('position_order')

        ingresos_list = []
        descuentos_list = []
        
        total_ingresos = Decimal("0.00")
        total_descuentos = Decimal("0.00")
        liquido_pagar = Decimal("0.00")

        for item in all_concepts:
            if item.data_source in ['TOTALINGRESOS', 'TOTALDSCTO', 'LIQUIDOPAGAR']:
                if item.data_source == 'TOTALINGRESOS': total_ingresos = item.amount
                if item.data_source == 'TOTALDSCTO': total_descuentos = item.amount
                if item.data_source == 'LIQUIDOPAGAR': liquido_pagar = item.amount
                continue
            
            concept_data = {
                "code": item.position_order,
                "name": item.concept,
                "amount": item.amount
            }

            if item.payroll_type == 'INGRESOS':
                ingresos_list.append(concept_data)
            elif item.payroll_type == 'DESCUENTOS':
                descuentos_list.append(concept_data)
                
        work_details = getattr(payslip_owner_profile, 'work_details', None)
        issue_month_name = MONTHS_ES[target_date.month - 1]
        issue_date_es = f"{issue_month_name} {target_date.year}"
        print_date = timezone.now().strftime("%d/%m/%Y")
        aporte_patronal = round(total_ingresos * Decimal("0.09"), 2)

        payload = {
            "issue_date_es": issue_date_es,
            "print_date": print_date,
            "dni": payslip_owner_profile.dni,
            "first_name": payslip_owner_user.first_name,
            "last_name": payslip_owner_user.last_name,
            "position": payslip_owner_profile.position,
            "description": payslip_owner_profile.description,
            "condition": payslip_owner_profile.condition,
            "regimen": payslip_owner_profile.regimen,
            "category": payslip_owner_profile.category,
            "worked_days": work_details.worked_days if work_details else 0,
            "worked_hours": work_details.worked_hours if work_details else 0,
            "discount_lateness": ((work_details.discount_lateness or 0) + (work_details.personal_leave_hours or 0)) if work_details else 0,
            "start_date": payslip_owner_profile.start_date.strftime("%d/%m/%Y") if payslip_owner_profile.start_date else "—",
            "end_date": payslip_owner_profile.end_date.strftime("%d/%m/%Y") if payslip_owner_profile.end_date else "VIGENTE",
            "sistema_pension": payslip_owner_profile.descriptionSP or "—",
            "codigo_afiliado": payslip_owner_profile.identification_code or "—",
            "ingresos": ingresos_list,
            "descuentos": descuentos_list,
            "total_bruto": total_ingresos,
            "total_descuento": total_descuentos,
            "total_liquido": liquido_pagar,
            "aporte_patronal": aporte_patronal,
        }

        html = render_to_string('boleta.html', payload)
        pdf_buffer = BytesIO()
        pisa_status = pisa.CreatePDF(html, dest=pdf_buffer)

        if pisa_status.err:
            return Response(
                APIResponse.error(message="Error al generar el PDF."),
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        pdf_filename = f"boleta_{reference_payslip.id}.pdf"
        reference_payslip.pdf_file.save(pdf_filename, ContentFile(pdf_buffer.getvalue()))
        
        all_concepts.update(view_status='generated')

        pdf_url = request.build_absolute_uri(reference_payslip.pdf_file.url)
        qr_bytes = generate_qr_code(pdf_url)

        send_payslip_email(
            user=payslip_owner_user,
            secure_url=pdf_url,
            qr_bytes=qr_bytes,
            issue_date=reference_payslip.issue_date
        )

        description = f"Boleta generada del periodo {issue_date_es} para {payslip_owner_user.first_name}."
        AuditLog.objects.create(profile=user.profile, action="GENERAR BOLETA", description=description)

        return Response(
            APIResponse.success(
                data={
                    "id": str(reference_payslip.id),
                    "pdf_url": reference_payslip.pdf_file.url,
                    "view_status": 'generated'
                },
                message="Boleta generada exitosamente."
            ),
            status=status.HTTP_200_OK
        )